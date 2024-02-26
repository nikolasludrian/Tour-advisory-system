from openpyxl import *
workbook = load_workbook(filename=r'Table\路线表.xlsx')
sheet = workbook["Sheet1"]
scenics=sheet['B1:N1']
scenic_dict={}
i=0
for row in scenics: # 遍历每一行的单元格
    for column in row: # 遍历每一列的单元格
        scenic_dict[i]=column.value
        i=i+1
def read_info():
    xlsxPath = 'Table\景点表.xlsx'
    # 第一步打开工作簿
    wb = load_workbook(xlsxPath)
    # 第二步选取表单
    sheet = wb.active
    # 按行获取数据转换成列表
    rows_data = list(sheet.rows)
    # 获取表单的表头信息(第一行)，也就是列表的第一个元素
    titles = [title.value for title in rows_data[0]]
    #整个表格最终转换出来的字典数据列表
    all_row_dict = []
    # 遍历工作表的每一行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 检查每一行是否所有单元格都有内容
        if all(cell is not None for cell in row):
            # 如果这一行所有单元格都有内容，则将其添加到字典列表中
            row_dict = dict(zip(titles, row))
            all_row_dict.append(row_dict)
    for d in all_row_dict:
        d['dist']=-1       
    return all_row_dict     
def read_dist(flag):  #flag=0,1,2
    cell=sheet["B2:N14"]
    all_distance=[]
    for row in cell: # 遍历每一行的单元格
        current_list=[]
        for column in row: # 遍历每一列的单元格
            new_list=list(map(float,column.value[1:-1].split(',')))
            current_list.append(new_list[flag])
        all_distance.append(current_list) 
    return  all_distance
def quick_sort(alist, start, end):
    """快速排序"""
    if start >= end:  # 递归的退出条件
        return
    mide=alist[start]
    mid = mide['排序']  # 设定起始的基准元素
    low = start  # low为序列左边在开始位置的由左向右移动的游标
    high = end  # high为序列右边末尾位置的由右向左移动的游标
    while low < high:
        # 如果low与high未重合，high(右边)指向的元素大于等于基准元素，则high向左移动
        while low < high and alist[high]['排序'] >= mid:
            high -= 1
        alist[low] = alist[high]  # 走到此位置时high指向一个比基准元素小的元素,将high指向的元素放到low的位置上,此时high指向的位置空着,接下来移动low找到符合条件的元素放在此处
        # 如果low与high未重合，low指向的元素比基准元素小，则low向右移动
        while low < high and alist[low]['排序'] < mid:
            low += 1
        alist[high] = alist[low]  # 此时low指向一个比基准元素大的元素,将low指向的元素放到high空着的位置上,此时low指向的位置空着,之后进行下一次循环,将high找到符合条件的元素填到此处

    # 退出循环后，low与high重合，此时所指位置为基准元素的正确位置,左边的元素都比基准元素小,右边的元素都比基准元素大
    alist[low] = mide  # 将基准元素放到该位置,
    # 对基准元素左边的子序列进行快速排序
    quick_sort(alist, start, low - 1)  # start :0  low -1 原基准元素靠左边一位
    # 对基准元素右边的子序列进行快速排序
    quick_sort(alist, low + 1, end)  # low+1 : 原基准元素靠右一位  end: 最后
    return alist
def taxi_cost(dist):
    dist=round(dist)
    cost=0
    if dist<=2:
        cost=8
    elif 2<dist<=8:
        cost=8+1.8*(dist-2)
    elif 8<dist<=20:
        cost=18.8+2.7*(dist-8)
    else:
        cost=32.4+3.6*(dist-20)
    cost=round(cost)
    return cost
if __name__=='__main__':        
